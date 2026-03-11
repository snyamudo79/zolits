from __future__ import annotations
import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Tuple

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandParser
from django.db import transaction
from django.utils import timezone

from core.models import (
    Role,
    Region,
    Depot,
    Module,
    IssueSeverity,
    IssueStatus,
    Issue,
    UserProfile,
)

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # will raise if .xlsx is attempted without dependency

User = get_user_model()


class Command(BaseCommand):
    help = "Import issues from an Excel (.xlsx) or CSV file into the issues table"

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument("file", help="Path to .xlsx or .csv file")
        parser.add_argument(
            "--sheet",
            default="Issues",
            help="Worksheet name (for .xlsx only). Defaults to 'Issues'. If not found, uses active sheet.",
        )
        parser.add_argument(
            "--default-region",
            default="SOUTHERN",
            help="Fallback region name if none found in source columns",
        )
        parser.add_argument(
            "--only-empty-description",
            action="store_true",
            help="Import only rows where description is empty but other information exists",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).resolve()
        if not file_path.exists():
            raise SystemExit(f"File not found: {file_path}")

        # Ensure reference data exists
        self._ensure_reference_data()
        importer_user = self._ensure_import_user()

        rows = self._read_rows(file_path, sheet_name=options.get("sheet"))
        if not rows:
            self.stdout.write(self.style.WARNING("No rows found to import"))
            return

        # Build caches for quick lookups
        region_cache: Dict[str, Region] = {}
        depot_cache: Dict[Tuple[int, str], Depot] = {}  # (region_id, depot_name)
        module_cache: Dict[str, Module] = {}
        severity_medium = IssueSeverity.objects.get(name__iexact="MEDIUM")
        status_pending = IssueStatus.objects.get(name__iexact="PENDING")
        status_resolved = IssueStatus.objects.get(name__iexact="RESOLVED")

        prefix_counters: Dict[str, int] = self._compute_existing_prefix_counters()

        created = 0
        skipped = 0

        for src in rows:
            norm = {k.strip().lower(): (v if not isinstance(v, str) else v.strip()) for k, v in src.items()}

            def gv(keys):
                for k in keys:
                    v = norm.get(k)
                    if v not in (None, ""):
                        return v
                return None

            module_name = self._upper(
                gv(["issue type", "module"]) or "ALL"
            )
            description = self._upper(
                gv(["issue description", "description", "issue desc", "desc"]) or ""
            )
            resolver_name = self._upper(
                gv(["resolver", "resolved by", "issue resolved by", "resolver name"]) or ""
            )
            raised_val = gv(["received", "date received", "date issue raised", "date raised", "raised", "received date"])
            resolved_val = gv(["resolution date", "date resolved", "resolved date"])
            resolution_val = gv(["resolution", "resolution notes", "notes", "comment", "comments"])
            has_info = any(
                [
                    bool(module_name and module_name != "ALL"),
                    bool(description),
                    bool(raised_val),
                    bool(resolver_name),
                    bool(resolved_val),
                    bool(resolution_val),
                ]
            )
            if options.get("only_empty_description"):
                if description:
                    continue
                if not has_info:
                    other_info = any(
                        [
                            bool(v) and str(v).strip() != ""
                            for key, v in norm.items()
                            if key not in ("issue description", "description")
                        ]
                    )
                    if not other_info:
                        skipped += 1
                        continue
            else:
                if not description:
                    skipped += 1
                    continue

            region_name = self._upper(gv(["region"]) or options["default_region"])
            region = region_cache.get(region_name)
            if not region:
                region = self._get_or_create_region(region_name)
                region_cache[region_name] = region

            # Per instruction, do not take depot from source; use UNSPECIFIED in the region
            depot = depot_cache.get((region.id, "UNSPECIFIED"))
            if not depot:
                depot = Depot.objects.get_or_create(region=region, name="UNSPECIFIED")[0]
                depot_cache[(region.id, "UNSPECIFIED")] = depot

            module = module_cache.get(module_name)
            if not module:
                module = Module.objects.get_or_create(name=module_name)[0]
                module_cache[module_name] = module

            # Timestamps
            raised_dt = self._parse_datetime(norm.get("received"))
            resolved_dt = self._parse_datetime(norm.get("resolution date"))

            # Status based on presence of resolution date
            status = status_resolved if resolved_dt else status_pending

            # Resolver user (optional)
            resolver_user = self._get_or_create_user_by_name(resolver_name) if resolver_name else None

            # Resolution notes
            resolution_notes = self._upper(resolution_val or "")

            # Generate new issue_number per region prefix
            issue_number = self._next_issue_number(region, prefix_counters)

            with transaction.atomic():
                issue = Issue.objects.create(
                    issue_number=issue_number,
                    region=region,
                    depot=depot,
                    module=module,
                    functionality="",  # not provided by source mapping
                    description=description,
                    raised_by_name="",  # not provided by source mapping
                    contact_phone="",
                    issue_logged_by=importer_user,
                    assigned_to=None,
                    severity=severity_medium,
                    status=status,
                    date_issue_raised=raised_dt or timezone.now(),
                    resolution_notes=resolution_notes,
                    issue_resolved_by=resolver_user,
                    date_issue_resolved=resolved_dt,
                )
                created += 1

        self.stdout.write(self.style.SUCCESS(f"Imported {created} issues; skipped {skipped} incomplete rows"))

    # Helpers
    def _ensure_reference_data(self) -> None:
        IssueSeverity.objects.get_or_create(name="MEDIUM", defaults={"priority_order": 3})
        IssueStatus.objects.get_or_create(name="PENDING", defaults={"is_resolved_state": False})
        IssueStatus.objects.get_or_create(name="RESOLVED", defaults={"is_resolved_state": True})

    def _ensure_import_user(self) -> Any:
        # Use first superuser if available
        user = User.objects.filter(is_superuser=True).first()
        if user:
            # Ensure they have an ADMIN profile for consistency
            admin_role, _ = Role.objects.get_or_create(name="ADMIN")
            UserProfile.objects.get_or_create(user=user, defaults={"role": admin_role})
            return user

        # Otherwise create an importer admin
        admin_role, _ = Role.objects.get_or_create(name="ADMIN")
        user, created = User.objects.get_or_create(
            username="importer",
            defaults={"email": "importer@example.com", "is_staff": True, "is_superuser": True},
        )
        if created:
            user.set_password("ChangeMe123!")
            user.first_name = "Data"
            user.last_name = "Importer"
            user.save()
        UserProfile.objects.get_or_create(user=user, defaults={"role": admin_role})
        return user

    def _read_rows(self, file_path: Path, sheet_name: Optional[str]) -> Iterable[Dict[str, Any]]:
        if file_path.suffix.lower() == ".csv":
            with file_path.open("r", newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                return list(reader)

        if file_path.suffix.lower() in (".xlsx", ".xlsm"):
            if load_workbook is None:
                raise SystemExit(
                    "openpyxl is required to read .xlsx files. Install it or export your sheet to CSV."
                )
            wb = load_workbook(filename=str(file_path), data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            results = []
            for r in rows[1:]:
                results.append({headers[i]: r[i] for i in range(len(headers))})
            return results

        raise SystemExit(f"Unsupported file type: {file_path.suffix}")

    def _upper(self, v: Any) -> str:
        return str(v).upper().strip() if v is not None else ""

    def _parse_datetime(self, v: Any) -> Optional[timezone.datetime]:
        if v is None or v == "":
            return None
        # openpyxl already returns datetime for date cells
        if isinstance(v, timezone.datetime):
            return timezone.make_aware(v) if timezone.is_naive(v) else v
        if isinstance(v, timezone.date):
            dt = timezone.datetime.combine(v, timezone.datetime.min.time())
            return timezone.make_aware(dt)
        if isinstance(v, (int, float)):
            # Excel serials should already be converted by openpyxl; fallback to None
            return None
        if isinstance(v, str):
            v = v.strip()
            fmts = [
                "%d/%m/%y",
                "%d/%m/%Y",
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%m/%d/%Y",
                "%Y.%m.%d",
                "%d.%m.%Y",
                "%d %b %Y",
                "%d %B %Y",
                "%Y-%m-%d %H:%M",
                "%d/%m/%Y %H:%M",
            ]
            for fmt in fmts:
                try:
                    dt = timezone.datetime.strptime(v, fmt)
                    return timezone.make_aware(dt)
                except ValueError:
                    continue
        return None

    def _get_or_create_region(self, name: str) -> Region:
        # Preferred known codes
        code_map = {
            "SOUTHERN": "SU",
            "WESTERN": "WE",
            "EASTERN": "EA",
            "NORTHERN": "NO",
            "HARARE": "HA",
            "HEAD OFFICE": "HO",
        }
        # Try case-insensitive name match first
        existing = Region.objects.filter(name__iexact=name).first()
        if existing:
            return existing
        code = code_map.get(name, (re.sub(r"[^A-Z]", "", name)[:2] or "RG"))
        # If there's a region already using the same code, reuse it
        code_match = Region.objects.filter(code__iexact=code).first()
        if code_match:
            return code_match
        # Create new region with the computed code; handle race/uniqueness gracefully
        try:
            obj, _ = Region.objects.get_or_create(name=name, defaults={"code": code})
            return obj
        except Exception:
            fallback = Region.objects.filter(code__iexact=code).first()
            if fallback:
                return fallback
            # Last resort: create with a modified code
            n = 1
            while True:
                new_code = f"{code[:2]}{n}"
                if not Region.objects.filter(code__iexact=new_code).exists():
                    obj, _ = Region.objects.get_or_create(name=name, defaults={"code": new_code})
                    return obj
                n += 1

    def _get_or_create_user_by_name(self, full_name: str) -> Optional[Any]:
        full_name = full_name.strip()
        if not full_name:
            return None
        # Try to find an existing user by case-insensitive first+last name
        parts = [p for p in full_name.split() if p]
        qs = User.objects.all()
        if len(parts) >= 2:
            qs = qs.filter(first_name__iexact=parts[0], last_name__iexact=" ".join(parts[1:]))
            user = qs.first()
            if user:
                self._ensure_expert_profile(user)
                return user

        # Fallback by username
        username = re.sub(r"[^a-z0-9]+", ".", full_name.lower()).strip(".")
        user, created = User.objects.get_or_create(
            username=username,
            defaults={"email": f"{username}@example.com", "first_name": parts[0] if parts else full_name, "last_name": " ".join(parts[1:]) if len(parts) > 1 else ""},
        )
        if created:
            user.set_password("ChangeMe123!")
            user.save()
        self._ensure_expert_profile(user)
        return user

    def _ensure_expert_profile(self, user: Any) -> None:
        expert_role, _ = Role.objects.get_or_create(name="EXPERT")
        UserProfile.objects.get_or_create(user=user, defaults={"role": expert_role})

    def _compute_existing_prefix_counters(self) -> Dict[str, int]:
        counters: Dict[str, int] = {}
        for region in Region.objects.all():
            prefix = region.code.upper()
            # Find max numeric suffix for this prefix
            max_n = 0
            for s in Issue.objects.filter(issue_number__startswith=prefix).values_list("issue_number", flat=True):
                m = re.match(rf"^{re.escape(prefix)}(\d+)$", s or "")
                if m:
                    max_n = max(max_n, int(m.group(1)))
            counters[prefix] = max_n
        return counters

    def _next_issue_number(self, region: Region, counters: Dict[str, int]) -> str:
        prefix = region.code.upper()
        current = counters.get(prefix, 0) + 1
        counters[prefix] = current
        return f"{prefix}{current}"
